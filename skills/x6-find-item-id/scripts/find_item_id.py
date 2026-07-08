#!/usr/bin/env python3
"""Find X6 item IDs by keyword in item config workbooks and text exports."""

from __future__ import annotations

import argparse
import csv
import json
import sys
import warnings
from pathlib import Path
from typing import Iterable

field_limit = sys.maxsize
while True:
    try:
        csv.field_size_limit(field_limit)
        break
    except OverflowError:
        field_limit //= 10
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

IMPORTANT_COLUMNS = {
    "id",
    "display_name",
    "name",
    "type",
    "func_type",
    "func_target_id",
    "use_condition_id",
    "道具编号",
    "道具名称",
    "道具表唯一id",
    "道具使用功能类型",
    "参数",
    "类型",
}


def find_workspace(start: Path) -> Path:
    current = start.resolve()
    for path in [current, *current.parents]:
        if (path / "X6Game").is_dir():
            return path
    raise SystemExit("workspace root not found; pass --workspace <path>")


def item_roots(workspace: Path) -> list[Path]:
    roots = [
        workspace / "X6Game" / "DesignerConfigurations" / "item",
        workspace / "X6Game" / "DesignerConfigs" / "item",
    ]
    return [root for root in roots if root.exists()]


def stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact_record(headers: list[str], values: list[str]) -> dict[str, str]:
    record: dict[str, str] = {}
    for index, value in enumerate(values):
        if not value:
            continue
        header = headers[index] if index < len(headers) and headers[index] else f"col_{index + 1}"
        if header in IMPORTANT_COLUMNS or any(key in header for key in ("id", "名称", "类型", "功能", "参数")):
            record[header] = value
    if "id" not in record and len(values) >= 2 and values[1].isdigit():
        record["id"] = values[1]
    return record


def header_from_rows(rows: list[list[str]], row_index: int | None = None) -> list[str]:
    for row in rows[:10]:
        if "id" in row or any(cell in IMPORTANT_COLUMNS for cell in row):
            return row
    if row_index is None:
        row_index = len(rows) - 1
    for candidate in range(row_index - 1, max(-1, row_index - 6), -1):
        row = rows[candidate]
        if any(cell in IMPORTANT_COLUMNS for cell in row) or "id" in row:
            return row
    width = max((len(row) for row in rows[: row_index + 1]), default=0)
    return [f"col_{i + 1}" for i in range(width)]


def search_xlsx(path: Path, keyword: str) -> Iterable[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required to search .xlsx files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        rows = [[stringify(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
        headers = header_from_rows(rows)
        for row_number, row in enumerate(rows, start=1):
            matched_columns = [i for i, cell in enumerate(row) if keyword in cell]
            if not matched_columns:
                continue
            yield {
                "source": str(path),
                "sheet": sheet.title,
                "row": row_number,
                "matched_columns": [
                    headers[i] if i < len(headers) and headers[i] else f"col_{i + 1}"
                    for i in matched_columns
                ],
                "record": compact_record(headers, row),
            }


def search_txt(path: Path, keyword: str) -> Iterable[dict[str, object]]:
    try:
        text = path.read_text(encoding="utf-8-sig")
    except UnicodeDecodeError:
        text = path.read_text(encoding="gb18030", errors="replace")

    rows = list(csv.reader(text.splitlines()))
    headers = header_from_rows(rows)
    for row_number, row in enumerate(rows, start=1):
        row = [cell.strip() for cell in row]
        matched_columns = [i for i, cell in enumerate(row) if keyword in cell]
        if not matched_columns:
            continue
        yield {
            "source": str(path),
            "row": row_number,
            "matched_columns": [
                headers[i] if i < len(headers) and headers[i] else f"col_{i + 1}"
                for i in matched_columns
            ],
            "record": compact_record(headers, row),
        }


def iter_sources(workspace: Path) -> Iterable[Path]:
    for root in item_roots(workspace):
        yield from root.rglob("*.xlsx")
        yield from root.rglob("*.txt")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("keyword", help="item display name, item id, func_type, or other exact keyword")
    parser.add_argument("--workspace", type=Path, default=None, help="X6 workspace root")
    parser.add_argument("--limit", type=int, default=30, help="maximum matches to print")
    parser.add_argument("--json", action="store_true", help="print JSON instead of a compact text summary")
    args = parser.parse_args()

    workspace = args.workspace.resolve() if args.workspace else find_workspace(Path.cwd())
    matches: list[dict[str, object]] = []
    for source in iter_sources(workspace):
        try:
            found = search_xlsx(source, args.keyword) if source.suffix == ".xlsx" else search_txt(source, args.keyword)
            for match in found:
                matches.append(match)
                if len(matches) >= args.limit:
                    break
        except Exception as exc:  # Keep a corrupt workbook from hiding other useful hits.
            matches.append({"source": str(source), "error": str(exc)})
        if len(matches) >= args.limit:
            break

    if args.json:
        print(json.dumps(matches, ensure_ascii=False, indent=2))
        return 0 if matches else 1

    for match in matches:
        location = match["source"]
        if "sheet" in match:
            location += f" | {match['sheet']}"
        location += f" | row {match.get('row', '?')}"
        print(location)
        if "error" in match:
            print(f"  error: {match['error']}")
            continue
        print(f"  matched: {', '.join(match.get('matched_columns', []))}")
        record = match.get("record", {})
        if isinstance(record, dict):
            for key, value in record.items():
                print(f"  {key}: {value}")
    return 0 if matches else 1


if __name__ == "__main__":
    raise SystemExit(main())
